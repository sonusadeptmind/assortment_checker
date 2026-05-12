#!/usr/bin/env python3
"""
Search Assortment Relevance Evaluation Framework
-------------------------------------------------
Computes per-keyword Precision / Recall / F1 across model iterations,
persists QA labels, and generates a two-tab XLSX report.

Usage:
  python evaluate_iteration.py \
      --dataset       dataset.csv          \
      --new_iteration new_iteration.xlsx   \
      --iteration_num 2                    \
      --output_dir    outputs/             \
      --label_store   label_store.json
"""

import argparse
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from utils import (
    DEFAULT_MAX_AGE_DAYS,
    iter_catalog_records,
    parse_pid_list,
    safe_round,
)

SEP = "|"          # separator used when writing PID lists to cells
LABEL_TP = "TP"
LABEL_FP = "FP"


def pids_to_str(pids: set) -> str:
    """Serialise a set of PIDs into a sorted, pipe-separated string."""
    return SEP.join(sorted(str(p) for p in pids)) if pids else ""

def load_dataset(path: str) -> pd.DataFrame:
    """Load dataset.csv, normalise column names, and validate required columns."""
    df = pd.read_csv(path, dtype=str).fillna("")
    # Normalise column names; rename unnamed count columns so they are preserved
    new_cols, unnamed_idx = [], 0
    for c in df.columns:
        clean = c.strip().lower().replace(" ", "_").rstrip(")")
        if clean.startswith("unnamed"):
            clean = f"_count_{unnamed_idx}"
            unnamed_idx += 1
        new_cols.append(clean)
    df.columns = new_cols
    required = {"keyword", "prod_ids", "pids_to_remove", "pids_to_include", "results_editor_re"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"dataset.csv is missing columns: {missing}")
    return df


def load_new_iteration(path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Read the new-iteration XLSX into two DataFrames.

    The file is expected to have a ``Combined Final Data`` sheet (columns
    ``original_keyword``, ``product_ids``) and an optional ``Removed
    Candidates`` sheet (columns ``keyword``, ``removed_by``).  Missing
    sheets fall back to an empty DataFrame with the expected schema.

    Returns:
        Tuple of ``(combined_df, removed_df)``.
    """
    xl = pd.ExcelFile(path)
    combined = xl.parse("Combined Final Data", dtype=str).fillna("")
    combined.columns = [c.strip().lower().replace(" ", "_") for c in combined.columns]
    if "Removed Candidates" in xl.sheet_names:
        removed = xl.parse("Removed Candidates", dtype=str).fillna("")
        removed.columns = [c.strip().lower().replace(" ", "_") for c in removed.columns]
    else:
        removed = pd.DataFrame(columns=["keyword", "removed_by"])
    return combined, removed


def load_label_store(path: str) -> dict:
    """Load the persistent label store, or ``{}`` if the file is missing."""
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def save_label_store(store: dict, path: str) -> None:
    """Persist the label store dict to a JSON file."""
    with open(path, "w") as f:
        json.dump(store, f, indent=2)


def load_catalog(
    path: str,
    *,
    max_age_days: Optional[int] = DEFAULT_MAX_AGE_DAYS,
) -> dict[str, bool]:
    """Load ``catalog.jsonl`` into ``{product_id: in_stock}``.

    Each JSONL line is expected to look like::

        {"product_id": "...", "product_liveness": true|false, "updated_at": "...", ...}

    Records whose ``updated_at`` is missing or older than ``max_age_days``
    are skipped via :func:`iter_catalog_records`.
    """
    liveness: dict[str, bool] = {}
    for _line_num, obj in iter_catalog_records(path, max_age_days=max_age_days):
        pid = str(obj.get("product_id", "")).strip()
        if pid:
            liveness[pid] = bool(obj.get("product_liveness", True))
    return liveness


def load_iteration_history(path: str) -> list:
    """Load the iteration aggregate-history list, or ``[]`` if absent."""
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return []


def save_iteration_history(history: list, path: str) -> None:
    """Persist the iteration aggregate-history list to JSON on disk."""
    with open(path, "w") as f:
        json.dump(history, f, indent=2)

def seed_label_store(label_store: dict, keyword: str,
                     tps: set, fps: set, iteration: int):
    """
    Upsert QA labels into the label store from dataset.csv columns
    (pids_to_include → TP, pids_to_remove → FP).

    Labels ARE overwritten when the incoming data disagrees with the stored
    value — dataset.csv is treated as the authoritative source for each run.
    A conflict warning is printed whenever a confirmed label is flipped so that
    unintentional changes in dataset.csv do not go unnoticed.

    Conflict priority: if a PID appears in both `tps` and `fps` in the same
    call, FP wins (fps loop runs last).
    """
    kw_store = label_store.setdefault(keyword, {})

    for pid in tps:
        entry = kw_store.setdefault(pid, {"label": LABEL_TP,
                                          "iteration_first_seen": iteration,
                                          "source": "manual_qa"})
        if entry["label"] != LABEL_TP:
            print(f"  ⚠  label conflict [{keyword}] pid={pid}: "
                  f"{entry['label']} → {LABEL_TP} (dataset.csv overrides)")
            entry["label"] = LABEL_TP
        entry["iteration_last_confirmed"] = iteration

    for pid in fps:
        entry = kw_store.setdefault(pid, {"label": LABEL_FP,
                                          "iteration_first_seen": iteration,
                                          "source": "manual_qa"})
        if entry["label"] != LABEL_FP:
            print(f"  ⚠  label conflict [{keyword}] pid={pid}: "
                  f"{entry['label']} → {LABEL_FP} (dataset.csv overrides)")
            entry["label"] = LABEL_FP
        entry["iteration_last_confirmed"] = iteration

def evaluate_keyword(
    keyword:                str,
    current_prod_ids:       set,   # pinned + in-stock  (prev iteration)
    current_pids_to_include:set,   # known TPs from QA  (prev iteration)
    current_pids_to_remove: set,   # known FPs from QA  (prev iteration)
    current_re:             set,   # full RE mapping     (prev iteration)
    new_product_ids:        set,   # model output for this iteration
    label_store:            dict,
    iteration_num:          int,
    catalog:                dict,  # {product_id: bool} — True = in-stock
    traffic:                str = "",
) -> dict:
    """Evaluate a single keyword across the previous and new model iterations.

    Computes precision, recall, F1 (both standard and stock-adjusted),
    label coverage, TP retention, FP elimination, and regression signals.

    Args:
        keyword: The search keyword being evaluated.
        current_prod_ids: Pinned + in-stock product IDs from the previous iteration.
        current_pids_to_include: Known true-positive PIDs from prior QA.
        current_pids_to_remove: Known false-positive PIDs from prior QA.
        current_re: Full results-editor mapping from the previous iteration.
        new_product_ids: Model output product IDs for this iteration.
        label_store: Persistent keyword→PID→label mapping.
        iteration_num: Current 1-based iteration number.
        catalog: Product liveness map {product_id: bool} (True = in-stock).
        traffic: Optional traffic volume string for this keyword.

    Returns:
        Dict containing all computed metrics, alert strings, and updated
        dataset column values (prefixed with '_').
    """
    kw_labels = label_store.get(keyword, {})
    known_tps  = {pid for pid, m in kw_labels.items() if m["label"] == LABEL_TP}
    known_fps  = {pid for pid, m in kw_labels.items() if m["label"] == LABEL_FP}

    # Stock status (from catalog product_liveness)
    # A product is OOS when product_liveness is False in the catalog.
    # We check every PID that is relevant to this keyword evaluation.
    # Unknown PIDs (not in catalog) default to in-stock (True).
    all_relevant_pids = current_re | current_prod_ids | new_product_ids | known_tps | known_fps
    oos_pids = {pid for pid in all_relevant_pids if not catalog.get(pid, True)}

    # New RE = what the model proposes to pin
    new_re = new_product_ids.copy()

    # New prod_ids = new RE minus OOS products
    new_prod_ids = new_re - oos_pids

    # Label partitions on new results
    tp_in_new      = new_product_ids & known_tps
    fp_in_new      = new_product_ids & known_fps
    unknown_in_new = new_product_ids - known_tps - known_fps

    # Regression / improvement signals
    tp_dropped          = known_tps - new_product_ids
    tp_dropped_oos      = tp_dropped & oos_pids          # neutral  – OOS
    tp_dropped_in_stock = tp_dropped - oos_pids          # ⚠ bad   – regression

    fp_eliminated = current_pids_to_remove - new_product_ids   # ✅ good
    fp_remaining  = current_pids_to_remove & new_product_ids   # still present

    # pids_to_check
    # Newly added PIDs (not in previous RE or prod_ids) that have no QA label yet.
    # Also include previously present but never-labeled products.
    prev_all = current_re | current_prod_ids
    newly_added         = new_product_ids - prev_all
    unlabeled_existing  = (new_product_ids - newly_added) - known_tps - known_fps
    pids_to_check       = newly_added | unlabeled_existing

    # OOS-aware TP partitions (computed early; needed by precision gating)
    tp_oos              = known_tps & oos_pids
    available_tps       = known_tps - tp_oos       # in-stock known TPs
    tp_in_new_available = tp_in_new & available_tps

    # Bug 3 gate: empty-results + in-stock TPs → real failure, score as 0
    # When the model returns no products but there are in-stock TPs available,
    # this is a genuine failure (recall=0, precision=0) rather than an undefined
    # case.  Score as 0 so these keywords are included in aggregate averages and
    # do not silently deflate metrics by injecting zeros from keywords where no
    # measurement is possible (e.g. all TPs OOS or no TPs at all).
    empty_but_has_available_tps = (not new_product_ids) and (len(available_tps) > 0)

    # Metrics
    labeled_count    = len(tp_in_new) + len(fp_in_new)

    # labeled_precision: TP / (TP+FP) over labeled new results.
    # When empty_but_has_available_tps: 0 (model should have returned products).
    # When new results exist but none are labeled: None (no signal yet).
    labeled_precision = (
        len(tp_in_new) / labeled_count if labeled_count
        else (0.0 if empty_but_has_available_tps else None)
    )

    # Standard recall: TP retrieved / all known TPs (regardless of stock status).
    # Naturally 0 when known_tps exist but new_product_ids is empty.
    standard_recall = len(tp_in_new) / len(known_tps) if known_tps else None

    # Labeled F1: harmonic mean of labeled_precision and standard_recall.
    if labeled_precision is not None and standard_recall is not None:
        denom_lf1 = labeled_precision + standard_recall
        labeled_f1 = 2 * labeled_precision * standard_recall / denom_lf1 if denom_lf1 else 0.0
    else:
        labeled_f1 = None

    # Stock-adjusted recall: denominator excludes TPs that are currently OOS.
    # Numerator is also restricted to available TPs to keep recall in [0, 1].
    # (A TP that was OOS last iteration may re-appear in new_product_ids; counting
    # it in the numerator but not the denominator would push recall above 1.)
    stock_adj_recall = (
        len(tp_in_new_available) / len(available_tps) if available_tps else None
    )

    # Stock-adjusted precision: excludes OOS products from both numerator and
    # denominator so that stock status does not inflate or deflate precision.
    fp_in_new_available = fp_in_new - oos_pids
    stock_adj_labeled_count = len(tp_in_new_available) + len(fp_in_new_available)
    stock_adj_precision = (
        len(tp_in_new_available) / stock_adj_labeled_count if stock_adj_labeled_count
        else (0.0 if empty_but_has_available_tps else None)
    )

    # Stock-adjusted F1: harmonic mean of stock_adj_precision and stock_adj_recall
    if stock_adj_precision is not None and stock_adj_recall is not None:
        denom_sf1 = stock_adj_precision + stock_adj_recall
        stock_adj_f1 = 2 * stock_adj_precision * stock_adj_recall / denom_sf1 if denom_sf1 else 0.0
    else:
        stock_adj_f1 = None

    label_coverage = labeled_count / len(new_product_ids) if new_product_ids else None

    # tp_retention_rate: fraction of prev-iteration's confirmed TPs (known TPs
    # that were in the previous RE pinset) still present in new results.
    # Distinct from standard_recall which uses ALL ever-known TPs as the
    # denominator; this metric specifically measures short-term regression —
    # did we keep what we already had pinned?
    # Returns None when the previous RE contained no confirmed TPs.
    prev_re_tps  = current_re & known_tps
    tp_retained  = prev_re_tps & new_product_ids
    tp_retention = len(tp_retained) / len(prev_re_tps) if prev_re_tps else None

    fp_elim_rate = (
        len(fp_eliminated) / len(current_pids_to_remove)
        if current_pids_to_remove else None
    )

    # Unlabeled flag: >30% of new results are unlabelled
    # Denominator is total new products (not labeled count) so the ratio is
    # directly comparable to label_coverage (flag triggers when coverage < 0.70).
    if new_product_ids:
        unlabeled_ratio = len(unknown_in_new) / len(new_product_ids)
        unlabeled_flag  = unlabeled_ratio > 0.30
    else:
        unlabeled_ratio = None
        unlabeled_flag  = False

    return {
        # identity
        "keyword":          keyword,
        "traffic":          traffic,
        "iteration":        iteration_num,

        # volume counts
        "prev_result_count":    len(current_prod_ids),
        "new_result_count":     len(new_prod_ids),
        "tp_count":             len(tp_in_new),
        "fp_count":             len(fp_in_new),
        "unknown_count":        len(unknown_in_new),
        "pids_to_check_count":  len(pids_to_check),
        "newly_added_count":    len(newly_added),

        # core metrics
        "labeled_precision":  safe_round(labeled_precision),
        "standard_recall":    safe_round(standard_recall),
        "labeled_f1":         safe_round(labeled_f1),
        "stock_adj_precision":safe_round(stock_adj_precision),
        "stock_adj_recall":   safe_round(stock_adj_recall),
        "stock_adj_f1":       safe_round(stock_adj_f1),
        "label_coverage":     safe_round(label_coverage),
        "tp_retention_rate":  safe_round(tp_retention),
        "fp_elimination_rate":safe_round(fp_elim_rate),

        # unlabeled flag
        "unlabeled_flag":       "⚠ YES" if unlabeled_flag else "",
        "unlabeled_ratio":      safe_round(unlabeled_ratio),

        # alert / detail strings
        "tp_dropped_in_stock":  pids_to_str(tp_dropped_in_stock),
        "tp_dropped_oos":       pids_to_str(tp_dropped_oos),
        "fp_eliminated":        pids_to_str(fp_eliminated),
        "fp_remaining":         pids_to_str(fp_remaining),

        # updated dataset columns (sets – serialised later)
        "_new_prod_ids":            new_prod_ids,
        "_updated_pids_to_include": tp_in_new,
        "_updated_pids_to_remove":  fp_in_new,
        "_new_re":                  new_re,
        "_pids_to_check":           pids_to_check,
    }

def compute_aggregate(results: list, iteration_num: int) -> dict:
    # Metric averages are computed only over keywords with manual_qa_status == True
    qa_results = [r for r in results if r.get("manual_qa_status") is True]

    def avg_metric(key):
        vals = [r[key] for r in qa_results if r.get(key) is not None]
        return safe_round(sum(vals) / len(vals)) if vals else None

    int_keys = [
        "prev_result_count", "new_result_count",
        "tp_count", "fp_count", "unknown_count",
        "pids_to_check_count", "newly_added_count",
    ]
    agg = {"keyword": "AGGREGATE", "traffic": "", "iteration": iteration_num}
    for k in int_keys:
        agg[k] = sum(r[k] for r in results)

    for k in ["labeled_precision", "standard_recall", "labeled_f1",
              "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
              "label_coverage", "tp_retention_rate", "fp_elimination_rate",
              "unlabeled_ratio"]:
        agg[k] = avg_metric(k)

    flagged = sum(1 for r in results if r.get("unlabeled_flag") == "⚠ YES")
    agg["unlabeled_flag"] = f"{flagged} of {len(results)} flagged"
    agg["manual_qa_status"] = f"{len(qa_results)} of {len(results)} QA'd"

    for k in ["tp_dropped_in_stock", "tp_dropped_oos", "fp_eliminated", "fp_remaining"]:
        agg[k] = ""

    return agg

# Colour palette
C_HEADER_SUMMARY = "1F4E79"   # dark blue
C_HEADER_DATASET = "375623"   # dark green
C_HEADER_TEXT    = "FFFFFF"
C_AGG_BG         = "D9E1F2"   # light blue
C_GREEN          = "C6EFCE"
C_YELLOW         = "FFEB9C"
C_RED_BG         = "FFC7CE"
C_RED_TEXT       = "9C0006"
C_CHECK_HEADER   = "BF8F00"   # amber
C_CHECK_CELL     = "FFF2CC"
C_ALT_ROW        = "F5F7FA"


def _fmt(hex_color, font_color=None, bold=False):
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    font = Font(color=font_color or "000000", bold=bold,
                name="Arial", size=10)
    return fill, font


def _autofit(ws, min_w=12, max_w=45):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        lengths = [len(str(ws.cell(r, col_idx).value or ""))
                   for r in range(1, ws.max_row + 1)]
        ws.column_dimensions[col_letter].width = min(max(max(lengths) + 3, min_w), max_w)


def _style_summary(ws, n_data_rows: int, col_names: list):
    # Header row
    fill_h, font_h = _fmt(C_HEADER_SUMMARY, C_HEADER_TEXT, bold=True)
    for cell in ws[1]:
        cell.fill, cell.font = fill_h, font_h
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column index maps
    col_map = {name: idx + 1 for idx, name in enumerate(col_names)}
    metric_cols = ["labeled_precision", "standard_recall", "labeled_f1",
                   "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
                   "label_coverage", "tp_retention_rate", "fp_elimination_rate"]
    alert_col       = col_map.get("tp_dropped_in_stock")
    flag_col        = col_map.get("unlabeled_flag")

    fill_g,  _    = _fmt(C_GREEN)
    fill_y,  _    = _fmt(C_YELLOW)
    fill_r,  _    = _fmt(C_RED_BG)
    fill_a,  f_a  = _fmt(C_AGG_BG, bold=True)

    for row_idx in range(2, n_data_rows + 3):   # +2 header, +1 agg row
        is_agg = ws.cell(row_idx, 1).value == "AGGREGATE"

        if is_agg:
            for cell in ws[row_idx]:
                cell.fill = fill_a
                cell.font = Font(name="Arial", size=10, bold=True)
            continue

        # Alternating row shading (light)
        if (row_idx % 2) == 0:
            alt_fill = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW,
                                   fill_type="solid")
            for cell in ws[row_idx]:
                if cell.fill.fill_type == "none":
                    cell.fill = alt_fill

        # Metric colour-coding
        for mc in metric_cols:
            if mc not in col_map:
                continue
            cell = ws.cell(row_idx, col_map[mc])
            try:
                v = float(cell.value)
                cell.fill = fill_g if v >= 0.75 else (fill_y if v >= 0.50 else fill_r)
            except (TypeError, ValueError):
                pass

        # Regression alert
        if alert_col:
            cell = ws.cell(row_idx, alert_col)
            if cell.value and str(cell.value).strip():
                cell.fill = fill_r
                cell.font = Font(name="Arial", size=10, bold=True, color=C_RED_TEXT)

        # Unlabeled flag highlight
        if flag_col:
            cell = ws.cell(row_idx, flag_col)
            if cell.value and "YES" in str(cell.value):
                cell.fill = fill_r
                cell.font = Font(name="Arial", size=10, bold=True, color=C_RED_TEXT)

    _autofit(ws)
    ws.freeze_panes = "A2"


def _style_dataset(ws, col_names: list):
    fill_h, font_h = _fmt(C_HEADER_DATASET, C_HEADER_TEXT, bold=True)
    for cell in ws[1]:
        cell.fill, cell.font = fill_h, font_h
        cell.alignment = Alignment(horizontal="center")

    check_col = next((i + 1 for i, c in enumerate(col_names)
                      if c == "pids_to_check"), None)
    flag_col_ds = next((i + 1 for i, c in enumerate(col_names)
                        if c == "unlabeled_flag"), None)

    if check_col:
        ws.cell(1, check_col).fill = PatternFill(
            start_color=C_CHECK_HEADER, end_color=C_CHECK_HEADER, fill_type="solid")
        ws.cell(1, check_col).font = Font(name="Arial", size=10, bold=True,
                                          color=C_HEADER_TEXT)
        fill_chk, _ = _fmt(C_CHECK_CELL)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, check_col)
            if cell.value and str(cell.value).strip():
                cell.fill = fill_chk

    # Unlabeled flag column — red highlight when flagged
    if flag_col_ds:
        fill_flag, _ = _fmt(C_RED_BG)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, flag_col_ds)
            if cell.value and "YES" in str(cell.value):
                cell.fill = fill_flag
                cell.font = Font(name="Arial", size=10, bold=True, color=C_RED_TEXT)

    alt_fill = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.fill.fill_type == "none":
                    cell.fill = alt_fill

    _autofit(ws)
    ws.freeze_panes = "A2"


def generate_report(results: list, iteration_num: int, output_dir: str,
                    prev_agg: dict = None) -> str:
    output_path = str(Path(output_dir) / f"iteration_{iteration_num}_report.xlsx")

    # Tab 1: Summary
    summary_col_order = [
        "keyword", "traffic", "iteration", "manual_qa_status",
        "prev_result_count", "new_result_count",
        "tp_count", "fp_count", "unknown_count",
        "pids_to_check_count", "newly_added_count",
        "labeled_precision", "standard_recall", "labeled_f1",
        "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
        "label_coverage", "tp_retention_rate", "fp_elimination_rate",
        "unlabeled_flag", "unlabeled_ratio",
        "tp_dropped_in_stock", "tp_dropped_oos",
        "fp_eliminated", "fp_remaining",
    ]

    agg = compute_aggregate(results, iteration_num)

    # Delta row vs previous iteration aggregate (if available)
    delta_rows = []
    if prev_agg:
        delta = {"keyword": f"DELTA vs iter {iteration_num - 1}",
                 "traffic": "", "iteration": ""}
        for k in ["labeled_precision", "standard_recall", "labeled_f1",
                  "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
                  "label_coverage", "tp_retention_rate", "fp_elimination_rate"]:
            curr = agg.get(k)
            prev = prev_agg.get(k)
            if curr is not None and prev is not None:
                delta[k] = safe_round(curr - prev)
            else:
                delta[k] = None
        for k in summary_col_order:
            if k not in delta:
                delta[k] = ""
        delta_rows.append(delta)

    summary_rows = results + [agg] + delta_rows
    summary_df   = pd.DataFrame(summary_rows)[summary_col_order]

    # Tab 2: Dataset
    dataset_col_order = [
        "keyword", "prod_ids", "pids_to_remove",
        "pids_to_include", "results_editor_re", "pids_to_check",
        "unlabeled_flag",
    ]
    dataset_rows = [{
        "keyword":          r["keyword"],
        "prod_ids":         pids_to_str(r["_new_prod_ids"]),
        "pids_to_remove":   pids_to_str(r["_updated_pids_to_remove"]),
        "pids_to_include":  pids_to_str(r["_updated_pids_to_include"]),
        "results_editor_re":pids_to_str(r["_new_re"]),
        "pids_to_check":    pids_to_str(r["_pids_to_check"]),
        "unlabeled_flag":   r["unlabeled_flag"],
    } for r in results]
    dataset_df = pd.DataFrame(dataset_rows)[dataset_col_order]

    # Write XLSX
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        dataset_df.to_excel(writer, sheet_name="Dataset", index=False)

    wb = load_workbook(output_path)
    _style_summary(wb["Summary"], len(results), summary_col_order)
    _style_dataset(wb["Dataset"], dataset_col_order)
    wb.save(output_path)

    return output_path

def print_terminal_summary(results: list, iteration_num: int, prev_agg: dict = None):
    agg = compute_aggregate(results, iteration_num)
    qa_results = [r for r in results if r.get("manual_qa_status") is True]

    def fmt(v):
        return f"{float(v):.2%}" if v is not None else "N/A"

    print(f"\n{'═' * 65}")
    print(f"  ITERATION {iteration_num} — AGGREGATE REPORT")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'═' * 65}")
    print(f"  Keywords evaluated      : {len(results)}")
    print(f"  Keywords QA'd (F1 base) : {len(qa_results)}")
    print(f"  Total products (new)    : {agg['new_result_count']}")
    print(f"  Products to QA-check    : {agg['pids_to_check_count']}")
    print(f"  Newly added PIDs        : {agg['newly_added_count']}")
    print(f"{'─' * 65}")
    print(f"  Avg Labeled Precision   : {fmt(agg['labeled_precision'])}")
    print(f"  Avg Standard Recall     : {fmt(agg['standard_recall'])}")
    print(f"  Avg Labeled F1          : {fmt(agg['labeled_f1'])}")
    print(f"  Avg Stock-Adj Precision : {fmt(agg['stock_adj_precision'])}")
    print(f"  Avg Stock-Adj Recall    : {fmt(agg['stock_adj_recall'])}")
    print(f"  Avg Stock-Adj F1        : {fmt(agg['stock_adj_f1'])}")
    print(f"  Avg Label Coverage      : {fmt(agg['label_coverage'])}")
    print(f"  Avg TP Retention Rate   : {fmt(agg['tp_retention_rate'])}")
    print(f"  Avg FP Elimination Rate : {fmt(agg['fp_elimination_rate'])}")

    if prev_agg:
        print(f"{'─' * 65}")
        print(f"  DELTA vs previous iteration:")
        for k, label in [("labeled_precision",  "Lbl Precision  "),
                          ("standard_recall",    "Std Recall     "),
                          ("labeled_f1",         "Lbl F1         "),
                          ("stock_adj_precision", "Adj Precision  "),
                          ("stock_adj_recall",   "Adj Recall     "),
                          ("stock_adj_f1",       "Adj F1         ")]:
            curr = agg.get(k)
            prev = prev_agg.get(k)
            if curr is not None and prev is not None:
                delta = curr - prev
                arrow = "▲" if delta > 0 else ("▼" if delta < 0 else "─")
                print(f"    {label}: {arrow} {delta:+.2%}")

    # Regressions
    regressions = [r for r in results
                   if r.get("tp_dropped_in_stock", "").strip()]
    if regressions:
        print(f"\n  ⚠  REGRESSIONS — {len(regressions)} keyword(s) with in-stock TPs dropped:")
        for r in sorted(regressions,
                        key=lambda x: len(x["tp_dropped_in_stock"].split(SEP)),
                        reverse=True)[:10]:
            count = len(r["tp_dropped_in_stock"].split(SEP))
            print(f"     • {r['keyword']:<35} {count} TP(s) dropped")

    # Unlabeled flag warnings
    flagged = [r for r in results if r.get("unlabeled_flag") == "⚠ YES"]
    if flagged:
        print(f"\n  🏷  UNLABELED FLAG — {len(flagged)} keyword(s) with >30% unlabeled-to-labeled ratio:")
        for r in sorted(flagged, key=lambda x: x["unknown_count"], reverse=True)[:10]:
            ratio_str = f"{float(r['unlabeled_ratio']):.0%}" if r.get('unlabeled_ratio') is not None else "N/A"
            print(f"     • {r['keyword']:<35} {r['unknown_count']} unknown / ratio {ratio_str}")

    # Top QA priority queue
    needs_review = sorted(
        [r for r in results if r["pids_to_check_count"] > 0],
        key=lambda r: r["pids_to_check_count"],
        reverse=True,
    )[:10]
    if needs_review:
        print(f"\n  📋 Top keywords by QA review queue size:")
        for r in needs_review:
            print(f"     • {r['keyword']:<35} {r['pids_to_check_count']} PIDs to check")

    print(f"{'═' * 65}\n")

def generate_keyword_breakdown_csv(results: list, iteration_num: int,
                                    output_dir: str) -> str:
    """Write a flat CSV with one row per keyword and all computed metrics."""
    csv_cols = [
        "keyword", "traffic", "iteration", "manual_qa_status",
        "prev_result_count", "new_result_count",
        "tp_count", "fp_count", "unknown_count",
        "pids_to_check_count", "newly_added_count",
        "labeled_precision", "standard_recall", "labeled_f1",
        "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
        "label_coverage", "tp_retention_rate", "fp_elimination_rate",
        "unlabeled_flag", "unlabeled_ratio",
        "tp_dropped_in_stock", "tp_dropped_oos",
        "fp_eliminated", "fp_remaining",
    ]
    rows = [{k: r.get(k, "") for k in csv_cols} for r in results]
    df = pd.DataFrame(rows, columns=csv_cols)
    csv_path = str(Path(output_dir) / f"iteration_{iteration_num}_keyword_breakdown.csv")
    df.to_csv(csv_path, index=False)
    return csv_path

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate search assortment iteration"
    )
    parser.add_argument("--dataset",       required=True,
                        help="Path to dataset.csv")
    parser.add_argument("--new_iteration", required=True,
                        help="Path to new_iteration.xlsx")
    parser.add_argument("--iteration_num", type=int, required=True,
                        help="Current iteration number (1-based)")
    parser.add_argument("--output_dir",    default="outputs",
                        help="Directory for output report files")
    parser.add_argument("--catalog",       required=True,
                        help="Path to catalog.jsonl (product_id + product_liveness)")
    parser.add_argument("--label_store",   default="label_store.json",
                        help="Path to persistent label store JSON")
    parser.add_argument("--history",       default="iteration_history.json",
                        help="Path to iteration aggregate history JSON")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # 1. Load inputs
    print(f"\n→ Loading data for iteration {args.iteration_num} …")
    dataset_df               = load_dataset(args.dataset)
    combined_df, removed_df  = load_new_iteration(args.new_iteration)
    catalog                  = load_catalog(args.catalog)
    label_store              = load_label_store(args.label_store)
    history                  = load_iteration_history(args.history)

    print(f"   dataset.csv         : {len(dataset_df)} keywords")
    print(f"   Combined Final Data : {len(combined_df)} rows")
    print(f"   Removed Candidates  : {len(removed_df)} rows")
    print(f"   catalog.jsonl       : {len(catalog)} products ({sum(1 for v in catalog.values() if not v)} OOS)")

    # 2. Build new-iteration lookup map
    new_iter_map: dict[str, set] = {}
    for _, row in combined_df.iterrows():
        kw   = str(row.get("original_keyword", "")).strip()
        pids = parse_pid_list(row.get("product_ids", ""))
        if kw:
            new_iter_map.setdefault(kw, set()).update(pids)

    traffic_map: dict[str, str] = {}
    if "traffic" in combined_df.columns:
        for _, row in combined_df.iterrows():
            kw = str(row.get("original_keyword", "")).strip()
            if kw and kw not in traffic_map:
                traffic_map[kw] = str(row.get("traffic", ""))

    # 3. Seed label store from current dataset QA annotations
    print("→ Seeding label store from dataset.csv …")
    seed_iter = max(args.iteration_num - 1, 1)
    for _, row in dataset_df.iterrows():
        kw  = str(row.get("keyword", "")).strip()
        tps = parse_pid_list(row.get("pids_to_include", ""))
        fps = parse_pid_list(row.get("pids_to_remove",  ""))
        if kw:
            seed_label_store(label_store, kw, tps, fps, seed_iter)

    # 4. Evaluate each keyword
    print("→ Evaluating keywords …")
    results = []

    ds_keywords = set(dataset_df["keyword"].str.strip().tolist())
    new_keywords = set(new_iter_map.keys())
    added_kws   = new_keywords - ds_keywords
    dropped_kws = ds_keywords  - new_keywords

    if added_kws:
        print(f"   ℹ  {len(added_kws)} new keyword(s) in this iteration")
    if dropped_kws:
        print(f"   ℹ  {len(dropped_kws)} keyword(s) not present in new iteration")

    # Evaluate keywords present in dataset
    for _, row in dataset_df.iterrows():
        kw = str(row.get("keyword", "")).strip()
        if not kw:
            continue
        result = evaluate_keyword(
            keyword                 = kw,
            current_prod_ids        = parse_pid_list(row.get("prod_ids",          "")),
            current_pids_to_include = parse_pid_list(row.get("pids_to_include",   "")),
            current_pids_to_remove  = parse_pid_list(row.get("pids_to_remove",    "")),
            current_re              = parse_pid_list(row.get("results_editor_re", "")),
            new_product_ids         = new_iter_map.get(kw, set()),
            label_store             = label_store,
            iteration_num           = args.iteration_num,
            catalog                 = catalog,
            traffic                 = traffic_map.get(kw, ""),
        )
        result["manual_qa_status"] = str(row.get("manual_qa_status", "")).strip().upper() == "TRUE"
        results.append(result)

    # Evaluate brand-new keywords (in new iteration but not yet in dataset)
    for kw in sorted(added_kws):
        result = evaluate_keyword(
            keyword                 = kw,
            current_prod_ids        = set(),
            current_pids_to_include = set(),
            current_pids_to_remove  = set(),
            current_re              = set(),
            new_product_ids         = new_iter_map.get(kw, set()),
            label_store             = label_store,
            iteration_num           = args.iteration_num,
            catalog                 = catalog,
            traffic                 = traffic_map.get(kw, ""),
        )
        result["manual_qa_status"] = False  # brand-new keyword, not yet QA'd
        results.append(result)

    # 5. Previous iteration aggregate (for delta)
    prev_agg = None
    if history:
        last = history[-1]
        if last.get("iteration") == args.iteration_num - 1:
            prev_agg = last

    # 6. Terminal summary
    print_terminal_summary(results, args.iteration_num, prev_agg)

    # 7. Generate XLSX report
    report_path = generate_report(
        results, args.iteration_num, args.output_dir, prev_agg
    )
    print(f"✅ Report saved : {report_path}")

    # 8. Generate keyword breakdown CSV
    csv_path = generate_keyword_breakdown_csv(
        results, args.iteration_num, args.output_dir
    )
    print(f"✅ Keyword CSV  : {csv_path}")

    # 9. Persist label store & history
    save_label_store(label_store, args.label_store)
    print(f"✅ Label store  : {args.label_store}")

    current_agg = compute_aggregate(results, args.iteration_num)
    history_entry = {
        "iteration":           args.iteration_num,
        "timestamp":           datetime.now().isoformat(),
        "labeled_precision":   current_agg.get("labeled_precision"),
        "standard_recall":     current_agg.get("standard_recall"),
        "labeled_f1":          current_agg.get("labeled_f1"),
        "stock_adj_precision": current_agg.get("stock_adj_precision"),
        "stock_adj_recall":    current_agg.get("stock_adj_recall"),
        "stock_adj_f1":        current_agg.get("stock_adj_f1"),
        "label_coverage":      current_agg.get("label_coverage"),
        "tp_retention_rate":   current_agg.get("tp_retention_rate"),
        "fp_elimination_rate": current_agg.get("fp_elimination_rate"),
        "keywords_evaluated":  len(results),
        "total_pids_to_check": current_agg.get("pids_to_check_count"),
    }
    history.append(history_entry)
    save_iteration_history(history, args.history)
    print(f"✅ History saved: {args.history}\n")


if __name__ == "__main__":
    main()
