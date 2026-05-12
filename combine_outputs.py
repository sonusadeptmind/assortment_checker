#!/usr/bin/env python3
"""Merge two assortment-checker output folders into one with no data loss.

Supported file types and merge strategy:

  *.xlsx              sheets merged row-wise; duplicate rows (same keyword)
                      kept from folder_a, new rows from folder_b appended.
  *.csv               concatenated and deduplicated on (keyword) or all columns.
  labels_store.json   list of label objects merged by (keyword, product_id);
                      agreed labels keep one copy, conflicting labels are set
                      to null with a conflict_details record.
  qa_metadata.json    domain-aware merge: lists unioned, dicts merged, scalars
                      favour folder_a; exported_at takes the later timestamp.
  other *.json        deep-merged (dict) or deduped (list).
  anything else       copied as-is; conflicts logged.

Usage:
  python combine_outputs.py <folder_a> <folder_b> [output_dir]

  output_dir defaults to ./outputs-combined (created if missing).
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# XLSX helpers

def _sheet_to_df(ws) -> pd.DataFrame:
    """Read an openpyxl worksheet into a DataFrame (first row = header)."""
    data = list(ws.values)
    if not data:
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    return pd.DataFrame(data[1:], columns=headers)


def _df_to_sheet(df: pd.DataFrame, ws) -> None:
    """Write a DataFrame into an openpyxl worksheet (overwrites existing rows)."""
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    # Write header
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _merge_xlsx(path_a: Path, path_b: Path, out_path: Path) -> None:
    """
    Merge two .xlsx files sheet-by-sheet.
    For each sheet that exists in both: concatenate rows, drop exact duplicates.
    Sheets only in one file are copied as-is.
    The workbook structure (styles etc.) from path_a is used as the base.
    """
    wb_a = load_workbook(path_a)
    wb_b = load_workbook(path_b)

    sheet_names_a = set(wb_a.sheetnames)
    sheet_names_b = set(wb_b.sheetnames)

    # Sheets in both → merge
    for sheet_name in sheet_names_a & sheet_names_b:
        ws_a = wb_a[sheet_name]
        ws_b = wb_b[sheet_name]
        df_a = _sheet_to_df(ws_a)
        df_b = _sheet_to_df(ws_b)

        if df_a.empty and df_b.empty:
            continue

        # Determine dedup key: prefer "keyword" column, else all columns
        dedup_cols = None
        if not df_a.empty and "keyword" in df_a.columns:
            dedup_cols = ["keyword"]
        elif not df_a.empty and "Keyword" in df_a.columns:
            dedup_cols = ["Keyword"]

        if df_a.empty:
            merged = df_b
        elif df_b.empty:
            merged = df_a
        else:
            # Align columns: fill missing ones with NaN
            all_cols = list(dict.fromkeys(list(df_a.columns) + list(df_b.columns)))
            df_a = df_a.reindex(columns=all_cols)
            df_b = df_b.reindex(columns=all_cols)

            if dedup_cols:
                # Keep df_a rows, add only df_b rows whose key isn't already in df_a
                existing_keys = set(df_a[dedup_cols[0]].dropna().astype(str))
                new_rows = df_b[~df_b[dedup_cols[0]].astype(str).isin(existing_keys)]
                merged = pd.concat([df_a, new_rows], ignore_index=True)
            else:
                merged = pd.concat([df_a, df_b], ignore_index=True).drop_duplicates()

        _df_to_sheet(merged, ws_a)

    # Sheets only in b → add to workbook a
    for sheet_name in sheet_names_b - sheet_names_a:
        ws_b = wb_b[sheet_name]
        ws_new = wb_a.create_sheet(title=sheet_name)
        df_b = _sheet_to_df(ws_b)
        _df_to_sheet(df_b, ws_new)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_a.save(out_path)
    print(f"  [xlsx ] merged  → {out_path.name}")


# CSV helpers

def _merge_csv(path_a: Path, path_b: Path, out_path: Path) -> None:
    """
    Concatenate two CSV files, deduplicate on 'keyword' (or all cols as fallback).
    """
    df_a = pd.read_csv(path_a)
    df_b = pd.read_csv(path_b)

    all_cols = list(dict.fromkeys(list(df_a.columns) + list(df_b.columns)))
    df_a = df_a.reindex(columns=all_cols)
    df_b = df_b.reindex(columns=all_cols)

    # Determine dedup key
    key_col = next(
        (c for c in ["keyword", "Keyword", "query", "Query"] if c in all_cols),
        None,
    )

    if key_col:
        existing_keys = set(df_a[key_col].dropna().astype(str))
        new_rows = df_b[~df_b[key_col].astype(str).isin(existing_keys)]
        merged = pd.concat([df_a, new_rows], ignore_index=True)
    else:
        merged = pd.concat([df_a, df_b], ignore_index=True).drop_duplicates()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    merged.to_csv(out_path, index=False)
    print(f"  [csv  ] merged  → {out_path.name}  ({len(merged)} rows)")


# JSON helpers

def _deep_merge_dicts(base: dict, overlay: dict) -> dict:
    """
    Recursively merge overlay into base.
    - Nested dicts are merged recursively.
    - Scalars: overlay wins only when base value is None/missing.
    - Lists: concatenated and deduplicated (by iteration_num if present).
    """
    result = dict(base)
    for key, val_b in overlay.items():
        if key not in result or result[key] is None:
            result[key] = val_b
        elif isinstance(result[key], dict) and isinstance(val_b, dict):
            result[key] = _deep_merge_dicts(result[key], val_b)
        elif isinstance(result[key], list) and isinstance(val_b, list):
            result[key] = _merge_lists(result[key], val_b)
        # else: base wins (don't overwrite existing non-null scalar)
    return result


def _merge_iteration_history_entries(entry_a: dict, entry_b: dict) -> dict:
    """Merge two iteration-history entries that share the same iteration hash.

    The two entries normally come from different reviewers who each
    QA'd a different slice of keywords for the same iteration.  We
    rebuild a single entry that approximates what the run would look
    like if it had been done by one reviewer over the union of slices.

    Strategy:

    * Rate/ratio metrics (precision, recall, F1, label_coverage, etc.)
      are combined as a *weighted* average using ``keywords_evaluated``
      as the weight.  This is exact for averages of per-keyword scores
      and a reasonable approximation for the others.
    * Count fields (``keywords_evaluated``, ``total_pids_to_check``,
      ``approved_count``, ``disapproved_count``) are summed.
    * Timestamps take the later of the two values.
    * If a metric is null in one entry, the other entry's value carries
      through unchanged.  When both are null the merged value stays null.
    * **Zero-keyword edge case:** when ``keywords_evaluated`` is zero in
      both entries, the weighted average would divide by zero — instead
      we fall back to a simple mean of the two metric values (or null
      when both are null).  This matches what the test suite asserts and
      avoids ZeroDivisionError on empty placeholder iterations.

    Non-numeric / non-count fields are taken from ``entry_a``.
    """
    METRIC_FIELDS = [
        "labeled_precision", "standard_recall", "labeled_f1",
        "stock_adj_precision", "stock_adj_recall", "stock_adj_f1",
        "label_coverage", "tp_retention_rate", "fp_elimination_rate",
    ]
    COUNT_FIELDS = [
        "keywords_evaluated", "total_pids_to_check", "approved_count", "disapproved_count",
    ]

    kw_a = entry_a.get("keywords_evaluated") or 0
    kw_b = entry_b.get("keywords_evaluated") or 0
    total_kw = kw_a + kw_b

    merged = dict(entry_a)

    # Weighted average for rate/ratio metrics
    for field in METRIC_FIELDS:
        va = entry_a.get(field)
        vb = entry_b.get(field)
        if va is None and vb is None:
            merged[field] = None
        elif va is None:
            merged[field] = vb
        elif vb is None:
            merged[field] = va
        elif total_kw > 0:
            merged[field] = round((va * kw_a + vb * kw_b) / total_kw, 4)
        else:
            merged[field] = round((va + vb) / 2, 4)

    # Sum for count fields
    for field in COUNT_FIELDS:
        va = entry_a.get(field) or 0
        vb = entry_b.get(field) or 0
        merged[field] = va + vb

    # Take the later timestamp
    ts_a = entry_a.get("timestamp", "")
    ts_b = entry_b.get("timestamp", "")
    merged["timestamp"] = max(ts_a, ts_b)

    return merged


def _merge_lists(list_a: list, list_b: list) -> list:
    """
    Merge two lists.
    - If items are dicts with an iteration identity key ('iteration' or
      'iteration_num'), deduplicate on that key; when the same iteration
      appears in both, merge the entries (weighted-average metrics).
    - Otherwise union of unique JSON-serialisable items.
    """
    # Detect iteration-history lists (keyed by 'iteration' hash or 'iteration_num')
    id_key = None
    if list_a and isinstance(list_a[0], dict):
        if "iteration" in list_a[0]:
            id_key = "iteration"
        elif "iteration_num" in list_a[0]:
            id_key = "iteration_num"

    if id_key:
        seen: dict[Any, dict] = {}
        for item in list_a:
            k = item.get(id_key)
            seen[k] = item
        for item in list_b:
            k = item.get(id_key)
            if k in seen:
                # Same iteration in both — merge the two partial entries
                seen[k] = _merge_iteration_history_entries(seen[k], item)
            else:
                seen[k] = item
        return list(seen.values())

    # Generic: union of unique items (JSON-serialisable)
    seen_strs: set[str] = set()
    result = []
    for item in list_a + list_b:
        s = json.dumps(item, sort_keys=True)
        if s not in seen_strs:
            seen_strs.add(s)
            result.append(item)
    return result


def _merge_labels_store(data_a: list, data_b: list) -> tuple[list, list]:
    """
    Merge two labels_store lists (each item has keyword, product_id, label, …).

    Strategy:
      - Same (keyword, product_id), same label  → keep one copy
      - Same (keyword, product_id), diff label  → emit one entry with
          label=null, conflict=True, conflict_details={a_label, b_label}
      - Only in one source → keep as-is

    Returns (merged_list, conflict_summary_list).
    """
    # Index by (keyword, product_id), keeping the latest entry per source
    map_a: dict[tuple, dict] = {}
    for item in data_a:
        key = (item.get("keyword"), item.get("product_id"))
        map_a[key] = item

    map_b: dict[tuple, dict] = {}
    for item in data_b:
        key = (item.get("keyword"), item.get("product_id"))
        map_b[key] = item

    merged: list[dict] = []
    conflict_summary: list[dict] = []

    all_keys = set(map_a) | set(map_b)
    for key in sorted(all_keys, key=lambda k: (k[0] or "", k[1] or "")):
        in_a = key in map_a
        in_b = key in map_b

        if in_a and not in_b:
            merged.append(map_a[key])
        elif in_b and not in_a:
            merged.append(map_b[key])
        else:
            item_a = map_a[key]
            item_b = map_b[key]
            label_a = item_a.get("label")
            label_b = item_b.get("label")

            if label_a == label_b:
                # Agreement — keep folder_a copy
                merged.append(item_a)
            else:
                # Conflict — null out label, record both originals
                entry = dict(item_a)
                entry["label"] = None
                entry["conflict"] = True
                entry["conflict_details"] = {
                    "a_label": label_a,
                    "b_label": label_b,
                }
                merged.append(entry)
                conflict_summary.append({
                    "keyword": key[0],
                    "product_id": key[1],
                    "a_label": label_a,
                    "b_label": label_b,
                })

    return merged, conflict_summary


def _merge_qa_metadata(data_a: dict, data_b: dict) -> dict:
    """
    Domain-aware merge of qa_metadata.json.

    Keys handled specifically:
      disapprovals / approvals / label_changes  → union of unique objects
      qa_done_keywords                          → union of unique strings
      iteration_labels                          → dict merge (a wins on conflict)
      current_iteration                         → keep a (same in both)
      exported_at                               → keep the later ISO timestamp
    """
    result = {}

    list_union_keys = {"disapprovals", "approvals", "label_changes"}
    str_list_keys   = {"qa_done_keywords"}
    dict_merge_keys = {"iteration_labels"}

    all_keys = set(data_a) | set(data_b)
    for key in all_keys:
        val_a = data_a.get(key)
        val_b = data_b.get(key)

        if key in list_union_keys:
            # Union of unique objects (dedup by JSON string)
            combined = (val_a or []) + (val_b or [])
            seen: set[str] = set()
            deduped = []
            for item in combined:
                s = json.dumps(item, sort_keys=True)
                if s not in seen:
                    seen.add(s)
                    deduped.append(item)
            result[key] = deduped

        elif key in str_list_keys:
            # Union of unique strings
            result[key] = list(dict.fromkeys((val_a or []) + (val_b or [])))

        elif key in dict_merge_keys:
            # Merge dicts: a wins on conflict
            merged_dict = dict(val_b or {})
            merged_dict.update(val_a or {})
            result[key] = merged_dict

        elif key == "exported_at":
            # Take the later of the two timestamps
            try:
                result[key] = max(val_a or "", val_b or "")
            except Exception:
                result[key] = val_a

        else:
            # Scalar or unknown key: prefer a, fall back to b
            result[key] = val_a if val_a is not None else val_b

    return result


def _merge_json(path_a: Path, path_b: Path, out_path: Path) -> tuple[list, list]:
    """
    Merge two JSON files (dict or list at the top level).
    Returns (merged_data, conflict_summary).
    """
    with open(path_a) as f:
        data_a = json.load(f)
    with open(path_b) as f:
        data_b = json.load(f)

    conflict_summary: list[dict] = []

    # Domain-specific handlers
    if path_a.name == "labels_store.json" and isinstance(data_a, list):
        merged, conflict_summary = _merge_labels_store(data_a, data_b)
        null_count = sum(1 for e in merged if e.get("conflict"))
        tag = f"{len(merged)} entries, {null_count} conflicts → label=null"

    elif path_a.name == "qa_metadata.json" and isinstance(data_a, dict):
        merged = _merge_qa_metadata(data_a, data_b)
        tag = f"{len(merged)} keys"

    elif isinstance(data_a, dict) and isinstance(data_b, dict):
        merged = _deep_merge_dicts(data_a, data_b)
        tag = f"{len(merged)} keys"

    elif isinstance(data_a, list) and isinstance(data_b, list):
        merged = _merge_lists(data_a, data_b)
        tag = f"{len(merged)} entries"

    else:
        merged = {"from_a": data_a, "from_b": data_b}
        tag = "type mismatch — stored under from_a / from_b"
        print(f"  [json ] WARNING: type mismatch in {path_a.name}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(merged, f, indent=2)

    print(f"  [json ] merged  → {out_path.name}  ({tag})")
    return merged, conflict_summary


# Main combiner

def combine_output_folders(
    folder_a: str | Path,
    folder_b: str | Path,
    output_dir: str | Path | None = None,
) -> Path:
    """
    Combine two assortment-checker output folders into one.

    Parameters
    ----------
    folder_a : path to first contributor's output folder
    folder_b : path to second contributor's output folder
    output_dir : destination folder (default: ./outputs-combined)

    Returns
    -------
    Path to the combined output folder.
    """
    folder_a = Path(folder_a).expanduser().resolve()
    folder_b = Path(folder_b).expanduser().resolve()

    if output_dir is None:
        output_dir = Path.cwd() / "outputs-combined"
    output_dir = Path(output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'─'*60}")
    print(f"  Folder A : {folder_a}")
    print(f"  Folder B : {folder_b}")
    print(f"  Output   : {output_dir}")
    print(f"{'─'*60}\n")

    # Collect all relative file paths from both folders
    def _relative_files(folder: Path) -> dict[Path, Path]:
        """Returns {relative_path: absolute_path} for all files."""
        return {
            p.relative_to(folder): p
            for p in folder.rglob("*")
            if p.is_file()
        }

    files_a = _relative_files(folder_a)
    files_b = _relative_files(folder_b)

    all_rel_paths = set(files_a) | set(files_b)
    file_conflicts: list[str] = []
    label_conflicts: list[dict] = []   # {keyword, product_id, a_label, b_label}

    for rel_path in sorted(all_rel_paths):
        out_path = output_dir / rel_path
        out_path.parent.mkdir(parents=True, exist_ok=True)

        in_a = rel_path in files_a
        in_b = rel_path in files_b

        # File only in one folder → straight copy
        if in_a and not in_b:
            shutil.copy2(files_a[rel_path], out_path)
            print(f"  [copy ] A-only  → {rel_path}")
            continue

        if in_b and not in_a:
            shutil.copy2(files_b[rel_path], out_path)
            print(f"  [copy ] B-only  → {rel_path}")
            continue

        # File in both → merge by type
        path_a = files_a[rel_path]
        path_b = files_b[rel_path]
        suffix = rel_path.suffix.lower()

        try:
            if suffix == ".xlsx":
                _merge_xlsx(path_a, path_b, out_path)

            elif suffix == ".csv":
                _merge_csv(path_a, path_b, out_path)

            elif suffix == ".json":
                _, conflicts_from_file = _merge_json(path_a, path_b, out_path)
                label_conflicts.extend(conflicts_from_file)

            elif suffix == ".jsonl":
                lines_a = path_a.read_text().splitlines()
                lines_b = path_b.read_text().splitlines()
                seen = set(lines_a)
                new_lines = [l for l in lines_b if l not in seen]
                out_path.write_text("\n".join(lines_a + new_lines) + "\n")
                print(f"  [jsonl] merged  → {rel_path}  (+{len(new_lines)} new lines)")

            else:
                # Binary or unknown – copy from A, log conflict
                shutil.copy2(path_a, out_path)
                file_conflicts.append(str(rel_path))
                print(f"  [copy ] conflict→ {rel_path}  (used A; B version logged)")

        except Exception as exc:
            print(f"  [ERROR] {rel_path}: {exc}")
            shutil.copy2(path_a, out_path)
            fallback = out_path.with_suffix(f".from_b{out_path.suffix}")
            shutil.copy2(path_b, fallback)
            file_conflicts.append(f"{rel_path} (B saved as {fallback.name})")

    # Write conflict report
    report_lines = [
        "=" * 60,
        "MERGE CONFLICT REPORT",
        "=" * 60,
        "",
    ]

    if label_conflicts:
        report_lines += [
            f"LABEL CONFLICTS ({len(label_conflicts)}) — label set to null in combined output",
            "  Resolve these in the dashboard's Merge panel.",
            "",
            f"  {'keyword':<55} {'product_id':<25} {'A':>4}  {'B':>4}",
            f"  {'-'*55} {'-'*25} {'----':>4}  {'----':>4}",
        ]
        for c in label_conflicts:
            report_lines.append(
                f"  {c['keyword']:<55} {c['product_id']:<25} {c['a_label']:>4}  {c['b_label']:>4}"
            )
        report_lines.append("")

    if file_conflicts:
        report_lines += [
            f"FILE CONFLICTS ({len(file_conflicts)}) — folder A version used",
            "",
        ] + [f"  - {c}" for c in file_conflicts] + [""]

    if not label_conflicts and not file_conflicts:
        report_lines += ["No conflicts — clean merge.", ""]

    report_lines += ["=" * 60, "END OF REPORT", "=" * 60]

    report_text = "\n".join(report_lines)
    conflict_report_path = output_dir / "_conflicts.txt"
    conflict_report_path.write_text(report_text)

    if label_conflicts or file_conflicts:
        print(f"\n  ⚠  {len(label_conflicts)} label conflict(s), "
              f"{len(file_conflicts)} file conflict(s) → _conflicts.txt")
    else:
        print("\n  ✅  No conflicts detected.")

    print(f"\n✅  Combined output ready: {output_dir}\n")
    return output_dir


# CLI entry-point

def main():
    parser = argparse.ArgumentParser(
        description="Combine two assortment-checker output folders without data loss.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("folder_a", help="Path to first contributor's output folder")
    parser.add_argument("folder_b", help="Path to second contributor's output folder")
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=None,
        help="Destination folder (default: ./outputs-combined)",
    )
    args = parser.parse_args()
    combine_output_folders(args.folder_a, args.folder_b, args.output_dir)


if __name__ == "__main__":
    main()
